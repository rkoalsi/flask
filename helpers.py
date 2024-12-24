import io, re, requests
import pandas as pd
from functools import lru_cache
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from env import org_id, PURCHASE_ORDER_URL, PURCHASE_URL, ITEM_URL, INVENTORY_URL, BOOKS_URL,clientId,clientSecret, grantType,inventory_refresh_token, books_refresh_token
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.mime.text import MIMEText

SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 587
SENDER_EMAIL = 'useremailaddy78@gmail.com'  # Use your email
SENDER_PASSWORD = 'nuymwywuhecelhho'  # Use your email app password


def validate_file(file) -> dict:
    """
    Checks if the given Excel file contains both 'PL' and 'CI' sheets.

    Args:
        file (BytesIO): In-memory file data (from form data).

    Returns:
        dict: Response indicating whether the required sheets are present or not.
    """
    try:
        # Load the workbook from the file
        wb = load_workbook(file)
        
        # Get all sheet names
        sheet_names = wb.sheetnames
        
        # Check if both 'PL' and 'CI' sheets are present
        if 'PL' not in sheet_names or 'CI' not in sheet_names:
            missing_sheets = []
            if 'PL' not in sheet_names:
                missing_sheets.append('PL')
            if 'CI' not in sheet_names:
                missing_sheets.append('CI')
            
            # Return an error response if any sheet is missing
            return {
                'status': 'error',
                'message': f"Missing sheets: {', '.join(missing_sheets)}"
            }

        # If both sheets are found, return a success response
        return {
            'status': 'success',
            'message': 'Both PL and CI sheets are present.'
        }

    except Exception as e:
        # Handle any other errors (e.g., invalid file format)
        return {
            'status': 'error',
            'message': f"An error occurred: {str(e)}"
        }


def send_email_with_attachments_in_memory(workbook, subject, body, filename, email):
    """Send email with multiple in-memory attachments."""
    msg = MIMEMultipart()
    msg['From'] = SENDER_EMAIL
    msg['To'] = email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    # Attach each in-memory workbook
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(workbook)
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename="{filename}.xlsx"')
    msg.attach(part)

    # Send the email
    try:
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()  # Encrypt connection
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.sendmail(SENDER_EMAIL, email, msg.as_string())
        server.quit()
        print(f"Email sent to {email} with in-memory attachments.")
    except Exception as e:
        print(f"Failed to send email: {e}")

def get_access_token(tkn: str):
    r = None
    access_token = ""
    if tkn == "inventory":
        r = requests.post(
            INVENTORY_URL.format(
                clientId=clientId,
                clientSecret=clientSecret,
                grantType=grantType,
                inventory_refresh_token=inventory_refresh_token,
            )
        )
    elif tkn == "books":
        r = requests.post(
            BOOKS_URL.format(
                clientId=clientId,
                clientSecret=clientSecret,
                grantType=grantType,
                books_refresh_token=books_refresh_token,
            )
        )
    else:
        print("missing token type")
        return
    access_token = str(r.json()["access_token"])
    print(f"Got {tkn.capitalize()} Access Token: {access_token[-4:]}")
    return access_token




access_token = get_access_token("books")
headers = {"Authorization": f"Zoho-oauthtoken {access_token}"}
company_name = "Pettingzoo"

def save_combined_sheet(matched_ci, unmatched_ci, matched_pl, unmatched_pl):
    """
    Saves four DataFrames to two sheets in a combined Excel file in memory.

    Args:
        matched_ci (pandas.DataFrame): DataFrame containing matched CI data.
        unmatched_ci (pandas.DataFrame): DataFrame containing unmatched CI data.
        matched_pl (pandas.DataFrame): DataFrame containing matched PL data.
        unmatched_pl (pandas.DataFrame): DataFrame containing unmatched PL data.

    Returns:
        bytes: The combined Excel file content in memory.
    """

    try:
        # Create a workbook
        wb = Workbook()

        # Sheet 1: Write matched and unmatched CI data
        ws1 = wb.active
        ws1.title = "CI Data"
        
        # Add "Matched CI" title
        ws1.append(["Matched CI"])
        for row in dataframe_to_rows(matched_ci, index=None, header=True):
            ws1.append(row)

        # Add a gap of two rows
        ws1.append([])
        ws1.append([])

        # Add "Unmatched CI" title
        ws1.append(["Unmatched CI"])
        for row in dataframe_to_rows(unmatched_ci, index=None, header=True):
            ws1.append(row)

        # Sheet 2: Write matched and unmatched PL data
        ws2 = wb.create_sheet(title="PL Data")

        # Add "Matched PL" title
        ws2.append(["Matched PL"])
        for row in dataframe_to_rows(matched_pl, index=None, header=True):
            ws2.append(row)

        # Add a gap of two rows
        ws2.append([])
        ws2.append([])

        # Add "Unmatched PL" title
        ws2.append(["Unmatched PL"])
        for row in dataframe_to_rows(unmatched_pl, index=None, header=True):
            ws2.append(row)

        # Create an in-memory buffer
        output_buffer = io.BytesIO()

        # Save the workbook to the buffer
        wb.save(output_buffer)

        # Reset the buffer position to the beginning
        output_buffer.seek(0)

        return output_buffer.getvalue()

    except Exception as e:
        print(f"Error saving combined sheet: {e}")
        return None

@lru_cache(maxsize=None)
def compare_strings(s1, s2):
    # remove whitespace, double spaces, hypens and brackets
    s1 = str(s1).replace(",", "").replace(" ", "").replace("--", "").casefold()
    s2 = str(s2).replace(",", "").replace(" ", "").replace("--", "").casefold()
    # compare strings
    if s1 == s2:
        return True
    else:
        return False

def extract_table_data(file_path, sheet_name, start_row=17):
    """
    Extracts table data from a given sheet using pandas.
    Assumes that the table starts after the specified start_row.
    """
    # Read the sheet starting from the specified row
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=start_row,  engine='openpyxl')
    # Drop rows that are entirely NaN
    df = df.dropna(how="all")

    # Optionally, reset the index
    df.reset_index(drop=True, inplace=True)

    return df

def is_valid_name(name):
    """
    Checks if the name is valid.
    A valid name is a string that does not resemble unwanted formats such as
    mostly uppercase letters, special characters, or alphanumeric codes.
    """
    # Exclude strings that are all uppercase or contain patterns like "B4321/D1"
    if not isinstance(name, str):
        return False
    if re.match(
        r"^[A-Z0-9/\-*\s()]+$", name
    ):  # Matches uppercase letters, digits, special chars
        return False
    return True

def get_purchase_orders(items):
    po = []
    p = requests.get(
        url=PURCHASE_URL.format(org_id=org_id, search_text=company_name, page=1),
        headers=headers,
    )
    po.extend([x for x in p.json()["purchaseorders"] if x["status"] != "draft"])
    has_more_pages = bool(p.json()["page_context"]["has_more_page"])
    if has_more_pages:
        print("Purchase orders have More Pages")
    else:
        print('No More Purchase Orders')
    found_items = []
    found_names = set()  # Use a set for quick lookups
    print('Processing POs')
    # Process each purchase order
    for purchase_order in po:
        po_id = purchase_order.get("purchaseorder_id")
        if not po_id:
            continue  # Skip if purchaseorder_id is missing

        # Fetch detailed purchase order items
        response = requests.get(
            url=PURCHASE_ORDER_URL.format(org_id=org_id, purchase_order_id=po_id),
            headers=headers,
        )
        purchase_order_items = (
            response.json().get("purchaseorder", {}).get("line_items", [])
        )

        # Check for matching items in line_items
        for item in items:
            for line_item in purchase_order_items:
                item_name = line_item.get("name")
                rate = line_item.get("rate")

                if (
                    compare_strings(item_name, item["name"])
                    and item_name not in found_names
                ):
                    found_items.append({"rate": rate, "name": item_name})
                    found_names.add(item_name)
                    break
    for item in items:
        if item["name"] not in found_names:
            found_items.append({"rate": 0, "name": item["name"]})
    return found_items

# def get_purchase_orders(items):
#     print('Fetching Purchase Orders')
#     po = []
#     page=1
#     print('Checking Further Purchase Orders')
#     # Loop until there are no more pages
#     while True:
#         # Make the API request
#         response = requests.get(
#         url=PURCHASE_URL.format(
#             org_id=org_id, search_text=company_name, page=page
#         ),
#         headers=headers,
#         )

#         # Parse the JSON response
#         data = response.json()

#         # Extract purchase orders that are not in 'draft' status
#         po.extend([x for x in data["purchaseorders"] if x["status"] != "draft"])

#         # Check if there are more pages
#         has_more_pages = data["page_context"]["has_more_page"]
#         if has_more_pages:
#             print(f"Processing page {page}, more pages available...")
#             page += 1  # Increment the page number
#         else:
#             print("No more pages.")
#             break  # Exit the loop
#     found_items = []
#     found_names = set()  # Use a set for quick lookups

#     # Process each purchase order
#     for purchase_order in po:
#         po_id = purchase_order.get("purchaseorder_id")
#         if not po_id:
#             continue  # Skip if purchaseorder_id is missing
#         # Fetch detailed purchase order items
#         response = requests.get(
#             url=PURCHASE_ORDER_URL.format(org_id=org_id, purchase_order_id=po_id),
#             headers=headers,
#         )
#         purchase_order_items = (
#             response.json().get("purchaseorder", {}).get("line_items", [])
#         )

#         # Check for matching items in line_items
#         for item in items:
#             for line_item in purchase_order_items:
#                 item_name = line_item.get("name")
#                 rate = line_item.get("rate")

#                 if (
#                     compare_strings(item_name, item["name"])
#                     and item_name not in found_names
#                 ):
#                     found_items.append({"rate": rate, "name": item_name})
#                     found_names.add(item_name)
#                     break
#         for item in items:
#             if item["name"] not in found_names:
#                 found_items.append({"rate": 0, "name": item["name"]})
#     return found_items

def process_upload(input_file, email):
    # Extract table data from both sheets
    input_file.seek(0)
    input_file = io.BytesIO(input_file.read())
    pl_sheet = extract_table_data(input_file, "PL")
    ci_sheet = extract_table_data(input_file, "CI", start_row=16)

    pl_data = [
        x.replace("\n", "").strip() for x in pl_sheet["DESCRIPTION"] if is_valid_name(x)
    ]

    ci_data = [
        {
            "name": row["DESCRIPTION"].replace("\n", "").strip(),
            "hsn": str(int(row["HSN "])),
            "price": row[" Unit Price"],
        }
        for _, row in ci_sheet.iterrows()
        if isinstance(row["DESCRIPTION"], str) and is_valid_name(row['DESCRIPTION'])
    ]
    # Check if data is matching with the data on zoho and print a list of all found items, and not found items
    matched_pl, unmatched_pl, matched_ci, unmatched_ci = [], [], [], []
    data = get_purchase_orders(ci_data)
    print('Processing PL Data')
    # fetch all items from PL sheet on zoho
    for item in pl_data:
        x = requests.get(
            url=ITEM_URL.format(org_id=org_id, search_text=item), headers=headers
        )
        product = x.json()["items"]
        if len(product) > 0:
            product = product[0]
            product_name = product.get("item_name")
            if compare_strings(item, product_name):
                matched_pl.append({"name": product_name})
            else:
                unmatched_pl.append({"name": product_name})
    print('Done Processing PL Data')
    print('Processing CI Data')
    for item in ci_data:
        name = item.get("name")
        print(name)
        code = item.get("hsn")
        price = item.get("price")
        x = requests.get(
            url=ITEM_URL.format(org_id=org_id, search_text=name), headers=headers
        )
        product = x.json()["items"]
        if len(product) > 0:
            product = product[0]
            product_name = product.get("item_name")
            product_code = product.get("hsn_or_sac")
            product_price = next(
                (
                    entry["rate"]
                    for entry in data
                    if compare_strings(entry["name"], product_name)
                ),
                0,
            )
            if (
                compare_strings(name, product_name)
                and compare_strings(code, product_code)
                and compare_strings(price, product_price)
            ):
                matched_ci.append(
                    {
                        "name": name,
                        "hsn": code,
                        "price": price,
                    }
                )
            else:
                reasons = []
                if not compare_strings(name, product_name):
                    reasons.append(f"Name {name} not matched with {product_name} ")
                if not compare_strings(code, product_code):
                    reasons.append(f"HSN {code} not matched with {product_code} ")
                if not compare_strings(price, product_price):
                    reasons.append(f"Price {price} not matched with {product_price} ")

                reason = "; ".join(reasons)
                unmatched_ci.append(
                    {
                        "name": name,
                        "hsn": code,
                        "price": price,
                        "reason": reason,
                    }
                )
        else:
            reason = f"{name} Not found in zoho"
            unmatched_ci.append(
                {
                    "name": name,
                    "hsn": code,
                    "price": price,
                    "reason": reason,
                }
            )

    print('Done Processing CI Data')

    matched_pl_df = pd.DataFrame(sorted(matched_pl, key=lambda x: str(x["name"])))
    unmatched_pl_df = pd.DataFrame(sorted(unmatched_pl, key=lambda x: str(x["name"])))
    matched_ci_df = pd.DataFrame(sorted(matched_ci, key=lambda x: int(x["hsn"])))
    unmatched_ci_df = pd.DataFrame(sorted(unmatched_ci, key=lambda x: int(x["hsn"])))

    workbook = save_combined_sheet(matched_ci_df, unmatched_ci_df, matched_pl_df, unmatched_pl_df)

    print("Saved to files")
    filename = "CI & PL Data"
    subject = "CI & PL Workbook"
    body = "Please find the attached CI and PL verification files."
    send_email_with_attachments_in_memory(workbook, subject, body, filename, email)
    return 1